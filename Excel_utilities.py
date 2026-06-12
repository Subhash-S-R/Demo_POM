import openpyxl
from openpyxl.styles import PatternFill

def rowNum(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row

def colNum(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column

def readData(file, sheetname, row, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row, col).value

def writeData(file, sheetname, row, col, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row, col).value = data
    workbook.save(file)

def redFill(file, sheetname, row, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    red_color = PatternFill(fill_type="solid", start_color='ff0000', end_color='ff0000')
    sheet.cell(row, col).fill = red_color
    workbook.save(file)

def greenFill(file, sheetname, row, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    green_color = PatternFill(fill_type="solid", start_color='00ff00', end_color='00ff00')
    sheet.cell(row, col).fill = green_color
    workbook.save(file)

def readLocators(file, sheetname):
    rows = rowNum(file, sheetname)
    out = {}
    for row in range(2, rows+1):
        out[readData(file, sheetname, row, 1)] = (readData(file, sheetname, row, 2), readData(file, sheetname, row, 3))
    return out

def readArgs(file, sheetname):
    cols = colNum(file, sheetname)
    out=[]
    for col in range(1, cols+1):
        out+=[readData(file, sheetname, 1, col)]
    print(', '.join(out))
    return ', '.join(out)

def read_all_data(file, sheetname):
    rows = rowNum(file, sheetname)
    cols = colNum(file, sheetname)
    out=[]
    for row in range(2, rows+1):
        tu = ()
        for col in range(1, cols+1):
            tu+=(readData(file, sheetname, row, col),)
        out+=[tu]
    print(out)
    return out
